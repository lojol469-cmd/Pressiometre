"""
PRESSIOMETRE IA - Logiciel portable d'analyse pressiométrique assisté par IA
Modèle KIBALI (Mistral-7B géophysique) + Streamlit + CUDA RTX 5090
"""

import os
import io
import re
import sys
import threading
import warnings
import datetime
import numpy as np
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
import streamlit as st
import openpyxl
from pathlib import Path

warnings.filterwarnings("ignore")

# ─── Chemins portables ────────────────────────────────────────────────────────
BASE_DIR   = Path(__file__).parent
MODEL_PATH = BASE_DIR / "kibali-final-merged"
ENV_PATH   = BASE_DIR / "environment"

# ─── Configuration Streamlit ──────────────────────────────────────────────────
st.set_page_config(
    page_title="Pressiomètre IA",
    page_icon="🌍",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─── CSS personnalisé ────────────────────────────────────────────────────────
st.markdown("""
<style>
    .main-header {
        background: linear-gradient(135deg, #1a1a2e 0%, #16213e 50%, #0f3460 100%);
        padding: 20px; border-radius: 12px; margin-bottom: 20px;
        text-align: center; color: white;
    }
    .main-header h1 { font-size: 2.2em; margin:0; letter-spacing:2px; }
    .main-header p  { color:#a8d8ea; margin:4px 0 0; font-size:1em; }
    .metric-card {
        background: #1e2a3a; border-radius: 10px; padding: 15px;
        border-left: 4px solid #00b4d8; margin: 5px 0;
    }
    .metric-card h4 { color: #90e0ef; margin:0 0 4px; font-size:0.85em; }
    .metric-card p  { color: #ffffff; font-size:1.4em; font-weight:bold; margin:0; }
    .anomaly-tag {
        background:#ff4d4d22; border:1px solid #ff4d4d;
        color:#ff6b6b; border-radius:5px; padding:3px 8px;
        font-size:0.8em; display:inline-block; margin:2px;
    }
    .ok-tag {
        background:#4dff9122; border:1px solid #4dff91;
        color:#4dff91; border-radius:5px; padding:3px 8px;
        font-size:0.8em; display:inline-block; margin:2px;
    }
    .stChatMessage { border-radius: 10px; }
    .kibali-badge {
        background: linear-gradient(90deg,#6c63ff,#4481eb);
        color:white; padding:3px 10px; border-radius:20px;
        font-size:0.8em; font-weight:bold;
    }
</style>
""", unsafe_allow_html=True)

# ─── Chargement du modèle IA (une seule fois en cache) ────────────────────────
@st.cache_resource(show_spinner="Chargement du modèle KIBALI (Mistral-7B géophysique)…")
def load_kibali():
    try:
        import torch
        from transformers import AutoModelForCausalLM, AutoTokenizer
        tokenizer = AutoTokenizer.from_pretrained(str(MODEL_PATH), local_files_only=True)
        model = AutoModelForCausalLM.from_pretrained(
            str(MODEL_PATH),
            torch_dtype=torch.float16,
            device_map="auto",
            local_files_only=True,
        )
        model.eval()
        return model, tokenizer, True
    except Exception as e:
        return None, None, str(e)

# ─── Parsing d'un fichier Excel pressiométrique ───────────────────────────────

METADATA_KEYS = [
    "Projet :", "Localisation :", "Ref Sondage :", "Ref Essai :",
    "Ref Sonde :", "Prof. de l'essai (m) :", "Ref étalonnage :",
    "Ref calibrage :", "Passe de forage (m) :", "Tech. Utilisée:",
    "Outil de forage :", "Pression diff (bar) :", "Type de tubulure :", "Date :",
]

def parse_sheet(ws) -> dict:
    """Parse une feuille pressiométrique -> dict avec meta + mesures."""
    meta = {}
    mesures = []
    in_data = False

    for row in ws.iter_rows(values_only=True):
        vals = list(row)
        if vals[0] == "DONNEES BRUTES":
            in_data = True
            continue
        if not in_data:
            key = str(vals[0]).strip() if vals[0] else None
            val = vals[1]
            if key:
                meta[key] = val
        else:
            # Ligne de données : [palier, None, V30, P60, V60, ...]
            if vals[0] is not None and isinstance(vals[0], (int, float)):
                palier = int(vals[0])
                v30 = vals[2] if len(vals) > 2 else None
                p60 = vals[3] if len(vals) > 3 else None
                v60 = vals[4] if len(vals) > 4 else None
                # Ignorer les valeurs marquées '_-10'
                if isinstance(v30, str) or isinstance(v60, str):
                    v30 = v60 = None
                if p60 is not None or v30 is not None:
                    mesures.append({
                        "palier": palier,
                        "V30_cm3": v30,
                        "P60_MPa": p60,
                        "V60_cm3": v60,
                    })
    return {"meta": meta, "mesures": mesures}


def load_excel(uploaded_file_or_path) -> dict:
    """Charge tous les essais d'un fichier Excel pressiométrique."""
    if isinstance(uploaded_file_or_path, (str, Path)):
        wb = openpyxl.load_workbook(uploaded_file_or_path, data_only=True)
    else:
        wb = openpyxl.load_workbook(io.BytesIO(uploaded_file_or_path.read()), data_only=True)

    results = {}
    calibrage = None
    etalonnages = {}

    for shname in wb.sheetnames:
        ws = wb[shname]
        parsed = parse_sheet(ws)
        name_clean = shname.strip()

        if name_clean.lower().startswith("calibr"):
            calibrage = parsed
        elif name_clean.lower().startswith("etalon") or name_clean.lower().startswith("étalon"):
            etalonnages[name_clean] = parsed
        else:
            # Extraire la profondeur depuis le nom de la feuille
            depth = extract_depth(name_clean, parsed["meta"])
            results[name_clean] = {**parsed, "depth_m": depth}

    return {
        "essais": results,
        "calibrage": calibrage,
        "etalonnages": etalonnages,
    }


def extract_depth(sheet_name: str, meta: dict) -> float | None:
    """Extrait la profondeur de l'essai depuis le nom de feuille ou les métadonnées."""
    # Depuis les métadonnées en priorité
    for k, v in meta.items():
        if "prof" in k.lower() and v is not None:
            try:
                s = str(v).replace(",", ".").replace("m", "").replace("M", "").strip()
                return float(s)
            except ValueError:
                pass
    # Depuis le nom de la feuille
    m = re.search(r"(\d+(?:[.,]\d+)?)\s*m", sheet_name, re.IGNORECASE)
    if m:
        return float(m.group(1).replace(",", "."))
    return None


# ─── Nettoyage dynamique (IA + règles physiques) ─────────────────────────────

def clean_essai(essai: dict, calibrage: dict | None, etalonnage: dict | None) -> pd.DataFrame:
    """
    Nettoie et corrige les données brutes d'un essai pressiométrique.
    Applique :
      1. Suppression des valeurs nulles / hors plage
      2. Correction de rigidité (calibrage)
      3. Correction de pression différentielle (étalonnage / pression_diff)
      4. Détection d'anomalies (rupture de membrane, points aberrants)
    """
    df = pd.DataFrame(essai["mesures"])
    if df.empty:
        return df

    df = df.dropna(subset=["P60_MPa"])
    df = df[df["P60_MPa"] >= 0]
    df = df[df["P60_MPa"] <= 20]  # max physique raisonnable

    # Correction volume – rigidité de sonde (calibrage)
    if calibrage and calibrage["mesures"]:
        df_cal = pd.DataFrame(calibrage["mesures"]).dropna(subset=["P60_MPa"])
        if not df_cal.empty:
            # Interpolation linéaire de la correction de volume
            from scipy.interpolate import interp1d
            try:
                f_v30 = interp1d(df_cal["P60_MPa"], df_cal["V30_cm3"].fillna(0),
                                 bounds_error=False, fill_value="extrapolate")  # type: ignore[call-overload]
                f_v60 = interp1d(df_cal["P60_MPa"], df_cal["V60_cm3"].fillna(0),
                                 bounds_error=False, fill_value="extrapolate")  # type: ignore[call-overload]
                df["V30_corr_cm3"] = df["V30_cm3"] - f_v30(df["P60_MPa"])
                df["V60_corr_cm3"] = df["V60_cm3"] - f_v60(df["P60_MPa"])
            except Exception:
                df["V30_corr_cm3"] = df["V30_cm3"]
                df["V60_corr_cm3"] = df["V60_cm3"]
        else:
            df["V30_corr_cm3"] = df["V30_cm3"]
            df["V60_corr_cm3"] = df["V60_cm3"]
    else:
        df["V30_corr_cm3"] = df["V30_cm3"]
        df["V60_corr_cm3"] = df["V60_cm3"]

    # Correction pression (pression différentielle + étalonnage)
    pression_diff_bar = 0.0
    try:
        val = essai["meta"].get("Pression diff (bar) :", 0) or 0
        pression_diff_bar = float(str(val).replace(",", "."))
    except (ValueError, TypeError):
        pass

    pression_diff_mpa = pression_diff_bar / 10.0

    if etalonnage and etalonnage["mesures"]:
        df_et = pd.DataFrame(etalonnage["mesures"]).dropna(subset=["P60_MPa"])
        if not df_et.empty:
            from scipy.interpolate import interp1d
            try:
                f_p = interp1d(df_et["P60_MPa"], df_et["P60_MPa"],
                               bounds_error=False, fill_value="extrapolate")  # type: ignore[call-overload]
                df["P60_corr_MPa"] = df["P60_MPa"] + pression_diff_mpa
            except Exception:
                df["P60_corr_MPa"] = df["P60_MPa"] + pression_diff_mpa
        else:
            df["P60_corr_MPa"] = df["P60_MPa"] + pression_diff_mpa
    else:
        df["P60_corr_MPa"] = df["P60_MPa"] + pression_diff_mpa

    df["P60_corr_MPa"] = df["P60_corr_MPa"].clip(lower=0)

    # Volume moyen corrigé
    df["Vm_corr_cm3"] = (df["V30_corr_cm3"].fillna(0) + df["V60_corr_cm3"].fillna(0)) / 2

    # Détection d'anomalies (rupture de membrane = saut de volume >200 cm3)
    df = df.sort_values("P60_corr_MPa").reset_index(drop=True)
    if "V60_corr_cm3" in df.columns and len(df) > 1:
        dv = df["V60_corr_cm3"].diff().abs()
        df["anomalie"] = dv > 200
    else:
        df["anomalie"] = False

    return df


# ─── Calcul des paramètres pressiométriques (NF P 94-110) ────────────────────

def compute_pressiometre_params(df: pd.DataFrame, depth_m: float) -> dict:
    """
    Calcule Pl* (pression limite nette), Em (module pressiométrique), Pf (pression de fluage)
    selon NF P 94-110 / méthode Ménard simplifiée.
    """
    if df.empty or len(df) < 3:
        return {}

    df_valid = df[~df["anomalie"]].copy()
    if df_valid.empty:
        return {}

    p = df_valid["P60_corr_MPa"].to_numpy(dtype=float)
    v = df_valid["Vm_corr_cm3"].to_numpy(dtype=float)

    # V0 : volume initial à pression nulle (interpolé)
    v0 = float(np.interp(0, p, v)) if len(p) > 1 else float(v[0])

    # Pression de fluage (Pf) : point d'inflexion sur la courbe P-V
    # Estimation par la dérivée seconde
    pf = None
    if len(p) >= 4:
        try:
            dv_dp = np.gradient(v, p)
            d2v_dp2 = np.gradient(dv_dp, p)
            idx_pf = np.argmax(np.abs(d2v_dp2[1:-1])) + 1
            pf = float(p[idx_pf])
        except Exception:
            pf = float(p[len(p) // 2]) if len(p) >= 2 else None

    # Pression limite (Pl) : pression pour laquelle V = 2*V0 + Vs (volume de la sonde)
    # Approximation : Pl estimée comme la pression au maximum ou extrapolée
    vs_cm3 = 535.0  # volume initial standard sonde Ménard (cm3)  
    v_limite = 2 * vs_cm3 + v0
    pl = None
    if v[-1] >= v_limite:
        pl = float(np.interp(v_limite, v, p))
    else:
        # Extrapolation linéaire sur les 3 derniers points
        if len(p) >= 3:
            try:
                coeffs = np.polyfit(v[-3:], p[-3:], 1)
                pl_ext = np.polyval(coeffs, v_limite)
                pl = float(pl_ext) if pl_ext > p[-1] else float(p[-1]) * 1.2
            except Exception:
                pl = float(p[-1]) * 1.2

    # Module pressiométrique Em = 2.66 * (1+v) * V0 * dP/dV
    # Prise sur la partie linéaire (entre 1/3 et 2/3 de la courbe)
    em = None
    if len(p) >= 4:
        try:
            i1 = max(1, len(p) // 3)
            i2 = min(len(p) - 1, 2 * len(p) // 3)
            dp = p[i2] - p[i1]
            dv = v[i2] - v[i1]
            if dv > 0:
                # coefficient de Poisson 0.33 (sol courant)
                em = 2.66 * v0 * (dp / dv)  # MPa
        except Exception:
            em = None

    # Classification du sol selon Em et Pl
    sol_type = classify_soil(em, pl)

    return {
        "depth_m": depth_m,
        "V0_cm3": round(v0, 1),
        "Pf_MPa": round(pf, 3) if pf else None,
        "Pl_MPa": round(pl, 3) if pl else None,
        "Em_MPa": round(em, 1) if em else None,
        "sol_type": sol_type,
        "n_paliers": len(df_valid),
        "anomalies": int(df["anomalie"].sum()),
    }


def classify_soil(em: float | None, pl: float | None) -> str:
    """Classification géotechnique basée sur Em et Pl (Ménard)."""
    if em is None or pl is None:
        return "Indéterminé"
    if pl < 0.3:
        return "Sol très mou (argile molle, tourbe)"
    elif pl < 1.0:
        if em < 5:
            return "Sol mou (argile peu consolidée)"
        else:
            return "Sol meuble (limon, sable lâche)"
    elif pl < 2.0:
        if em < 20:
            return "Sol de consistance moyenne"
        else:
            return "Sable compact / Limon dense"
    elif pl < 4.0:
        return "Sol raide (argile raide, sable dense)"
    elif pl < 8.0:
        return "Sol très raide / Gravier"
    else:
        return "Rocher altéré / Sol rocheux"


# ─── Interface KIBALI IA ──────────────────────────────────────────────────────

def ask_kibali(model, tokenizer, question: str, context: str = "") -> str:
    """Interroge le modèle KIBALI avec le contexte pressiométrique."""
    import torch
    system = (
        "Tu es KIBALI, un expert en géophysique et géotechnique spécialisé dans "
        "l'essai pressiométrique Ménard (NF P 94-110). Tu analyses des données "
        "pressiométriques (courbes P-V, module Em, pression limite Pl, fluage Pf), "
        "tu interprètes les anomalies et fournis des recommandations techniques "
        "claires en français."
    )
    if context:
        prompt = f"[INST] <<SYS>>\n{system}\n<</SYS>>\n\nContexte des données :\n{context}\n\nQuestion : {question} [/INST]"
    else:
        prompt = f"[INST] <<SYS>>\n{system}\n<</SYS>>\n\n{question} [/INST]"

    inputs = tokenizer(prompt, return_tensors="pt", truncation=True, max_length=2048).to(model.device)
    with torch.no_grad():
        outputs = model.generate(
            **inputs,
            max_new_tokens=512,
            temperature=0.7,
            top_p=0.9,
            do_sample=True,
            pad_token_id=tokenizer.eos_token_id,
        )
    full = tokenizer.decode(outputs[0], skip_special_tokens=True)
    # Extraire la réponse après [/INST]
    if "[/INST]" in full:
        return full.split("[/INST]", 1)[-1].strip()
    return full.strip()


# ─── Visualisations ───────────────────────────────────────────────────────────

def plot_courbe_pv(df: pd.DataFrame, depth_m: float, params: dict) -> go.Figure:
    """Trace la courbe pressiométrique P-V corrigée."""
    fig = go.Figure()

    df_valid = df[~df["anomalie"]]
    df_anom  = df[df["anomalie"]]

    fig.add_trace(go.Scatter(
        x=df_valid["Vm_corr_cm3"], y=df_valid["P60_corr_MPa"],
        mode="lines+markers", name="Courbe P-V",
        line=dict(color="#00b4d8", width=2),
        marker=dict(size=7, color="#90e0ef"),
    ))

    if not df_anom.empty:
        fig.add_trace(go.Scatter(
            x=df_anom["Vm_corr_cm3"], y=df_anom["P60_corr_MPa"],
            mode="markers", name="Anomalie",
            marker=dict(size=12, color="#ff4d4d", symbol="x"),
        ))

    # Ligne Pf
    if params.get("Pf_MPa"):
        fig.add_hline(y=params["Pf_MPa"], line_dash="dash",
                      line_color="#ffd166", annotation_text=f"Pf = {params['Pf_MPa']} MPa")

    # Ligne Pl
    if params.get("Pl_MPa"):
        fig.add_hline(y=params["Pl_MPa"], line_dash="dot",
                      line_color="#ef476f", annotation_text=f"Pl = {params['Pl_MPa']} MPa")

    fig.update_layout(
        title=f"Courbe Pressiométrique — Profondeur {depth_m} m",
        xaxis_title="Volume Vm corrigé (cm³)",
        yaxis_title="Pression P60 corrigée (MPa)",
        template="plotly_dark",
        height=420,
        showlegend=True,
        margin=dict(l=50, r=20, t=55, b=40),
    )
    return fig


def plot_profil(all_params: list[dict]) -> go.Figure:
    """Trace le profil géotechnique (Em, Pl vs profondeur)."""
    df_p = pd.DataFrame([p for p in all_params if p.get("depth_m") is not None])
    if df_p.empty:
        return go.Figure()

    df_p = df_p.sort_values("depth_m")
    fig = go.Figure()

    fig.add_trace(go.Scatter(
        x=df_p["Em_MPa"], y=df_p["depth_m"],
        mode="lines+markers", name="Em (MPa)",
        line=dict(color="#06d6a0", width=2),
        marker=dict(size=8),
    ))
    fig.add_trace(go.Scatter(
        x=df_p["Pl_MPa"], y=df_p["depth_m"],
        mode="lines+markers", name="Pl (MPa)",
        line=dict(color="#ef476f", width=2),
        marker=dict(size=8),
        xaxis="x2",
    ))

    fig.update_layout(
        title="Profil Géotechnique — Em et Pl vs Profondeur",
        yaxis=dict(title="Profondeur (m)", autorange="reversed"),
        xaxis=dict(title="Module Em (MPa)", domain=[0, 0.45], color="#06d6a0"),
        xaxis2=dict(title="Pression limite Pl (MPa)", domain=[0.55, 1.0],
                    anchor="y", color="#ef476f"),
        template="plotly_dark",
        height=500,
        margin=dict(l=60, r=60, t=60, b=40),
    )
    return fig


# ─── Application principale ───────────────────────────────────────────────────

def main():
    # En-tête
    st.markdown("""
    <div class="main-header">
        <h1>🌍 PRESSIOMÈTRE IA</h1>
        <p>Analyse pressiométrique intelligente · Modèle KIBALI (Mistral-7B Géophysique) · NF P 94-110</p>
    </div>
    """, unsafe_allow_html=True)

    # Sidebar
    with st.sidebar:
        st.markdown("### ⚙️ Configuration")
        use_ia = st.toggle("Activer l'IA KIBALI", value=True)
        st.markdown("---")
        st.markdown("### 📂 Charger des données")
        uploaded = st.file_uploader(
            "Fichier Excel pressiométrique",
            type=["xlsx"],
            accept_multiple_files=False,
        )

        # Fichiers pré-chargés
        st.markdown("**Fichiers disponibles :**")
        preset_files = {
            "Données pressio SP2": BASE_DIR / "Données pressio SP2.xlsx",
            "Relevé 20 premiers m": BASE_DIR / "releve des 20 premier metre.xlsx",
            "Relevé 30 premiers m": BASE_DIR / "releve des 30 premier metre.xlsx",
        }
        selected_preset = None
        for label, path in preset_files.items():
            if path.exists():
                if st.button(f"📊 {label}", use_container_width=True):
                    selected_preset = path

        st.markdown("---")
        st.markdown("### 🔬 À propos de KIBALI")
        st.markdown("""
        <span class='kibali-badge'>KIBALI</span>
        Mistral-7B fine-tuné sur la géophysique.
        Spécialisé : pressiomètre, ERT, GPR, géotechnique.
        """, unsafe_allow_html=True)

    # Chargement du modèle
    model, tokenizer, model_ok = None, None, False
    if use_ia:
        model, tokenizer, model_ok = load_kibali()
        if model_ok is True:
            st.sidebar.success("✅ KIBALI chargé (GPU)")
        else:
            st.sidebar.warning(f"⚠️ KIBALI non disponible : {model_ok}")

    # ─── Onglets principaux ───────────────────────────────────────────────────
    tab_data, tab_essai, tab_profil, tab_ia = st.tabs([
        "📋 Données brutes",
        "📈 Analyse essai",
        "🗻 Profil géo",
        "🤖 IA KIBALI",
    ])

    # Chargement des données
    data = None
    source_name = None

    if uploaded is not None:
        with st.spinner("Lecture du fichier…"):
            data = load_excel(uploaded)
        source_name = uploaded.name
    elif selected_preset:
        with st.spinner(f"Chargement de {selected_preset.name}…"):
            data = load_excel(selected_preset)
        source_name = selected_preset.name

    if data is None:
        st.info("👈 Chargez un fichier Excel pressiométrique depuis la barre latérale pour commencer.")

        # Afficher un exemple pédagogique
        with st.expander("📚 Comprendre l'essai pressiométrique Ménard", expanded=True):
            col1, col2 = st.columns(2)
            with col1:
                st.markdown("""
**Principe de l'essai**

L'essai pressiométrique Ménard (NF P 94-110) consiste à gonfler une sonde cylindrique 
dans un forage et à mesurer la déformation du sol sous pression contrôlée.

**Paramètres clés :**
- **V30, V60** : volumes lus à 30 s et 60 s (cm³)
- **P60** : pression appliquée à 60 s (MPa)
- **Em** : module pressiométrique (rigidité du sol)
- **Pf** : pression de fluage (fin de domaine pseudo-élastique)
- **Pl** : pression limite (résistance ultime du sol)

**Corrections :**
1. *Calibrage* : correction de la rigidité de la sonde (déformation mécanique)
2. *Étalonnage* : correction de pression hydrostatique et pertes de charge
3. *Pression diff* : correction de hauteur d'eau (colonne d'eau)
                """)
            with col2:
                st.markdown("""
**Classification des sols (Ménard) :**

| Type de sol | Em (MPa) | Pl (MPa) |
|---|---|---|
| Tourbe / argile molle | < 2 | < 0.3 |
| Argile peu consolidée | 2–5 | 0.3–1.0 |
| Limon, sable lâche | 5–15 | 0.5–1.5 |
| Sable compact | 10–30 | 1.5–3.0 |
| Argile raide, gravier | 20–80 | 2.0–5.0 |
| Sol rocheux / rocher altéré | > 80 | > 5.0 |

**Rapport Em/Pl :**
- Sols normalement consolidés : Em/Pl ≈ 7–12
- Sols surconsolidés : Em/Pl > 12
- Sols remaniés : Em/Pl < 7
                """)
        return

    # ─── ONGLET 1 : Données brutes ────────────────────────────────────────────
    with tab_data:
        st.subheader(f"📋 Fichier : {source_name}")
        essais = data["essais"]

        # Résumé du fichier
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.markdown(f'<div class="metric-card"><h4>Essais</h4><p>{len(essais)}</p></div>', unsafe_allow_html=True)
        with col2:
            has_cal = "✅" if data["calibrage"] else "❌"
            st.markdown(f'<div class="metric-card"><h4>Calibrage</h4><p>{has_cal}</p></div>', unsafe_allow_html=True)
        with col3:
            st.markdown(f'<div class="metric-card"><h4>Étalonnages</h4><p>{len(data["etalonnages"])}</p></div>', unsafe_allow_html=True)
        with col4:
            depths = [v["depth_m"] for v in essais.values() if v.get("depth_m") is not None]
            prof_str = f"{min(depths):.0f}–{max(depths):.0f} m" if depths else "N/A"
            st.markdown(f'<div class="metric-card"><h4>Profondeur</h4><p>{prof_str}</p></div>', unsafe_allow_html=True)

        st.markdown("---")

        # Tableau récapitulatif
        rows = []
        for name, ess in essais.items():
            n_mes = len([m for m in ess["mesures"] if m["P60_MPa"] is not None])
            rows.append({
                "Feuille": name,
                "Profondeur (m)": ess.get("depth_m"),
                "Sondage": ess["meta"].get("Ref Sondage :"),
                "Essai Réf": ess["meta"].get("Ref Essai :"),
                "Localisation": ess["meta"].get("Localisation :"),
                "Technique": ess["meta"].get("Tech. Utilisée:"),
                "N° paliers": n_mes,
                "Date": str(ess["meta"].get("Date :") or ""),
            })
        df_summary = pd.DataFrame(rows)
        st.dataframe(df_summary, use_container_width=True, height=350)

        # Données brutes d'un essai sélectionné
        st.markdown("#### 🔍 Détail données brutes")
        sel = st.selectbox("Sélectionner un essai", list(essais.keys()))
        if sel:
            df_raw = pd.DataFrame(essais[sel]["mesures"])
            if not df_raw.empty:
                st.dataframe(df_raw, use_container_width=True)
            else:
                st.warning("Aucune mesure valide dans cet essai.")

    # ─── ONGLET 2 : Analyse essai ─────────────────────────────────────────────
    with tab_essai:
        st.subheader("📈 Analyse d'un essai pressiométrique")
        essais = data["essais"]
        if not essais:
            st.warning("Aucun essai trouvé dans ce fichier.")
        else:
            sel2 = st.selectbox("Essai à analyser", list(essais.keys()), key="sel_essai")
            if sel2:
                ess = essais[sel2]
                depth = ess.get("depth_m", 0) or 0

                # Trouver le bon étalonnage
                ref_et = ess["meta"].get("Ref étalonnage :")
                etalonnage = None
                for k, v in data["etalonnages"].items():
                    if ref_et and ref_et in k:
                        etalonnage = v
                        break
                if etalonnage is None and data["etalonnages"]:
                    etalonnage = list(data["etalonnages"].values())[0]

                # Nettoyage
                with st.spinner("Nettoyage et correction des données…"):
                    df_clean = clean_essai(ess, data["calibrage"], etalonnage)

                # Paramètres
                params = compute_pressiometre_params(df_clean, depth)

                # Affichage métriques
                if params:
                    col1, col2, col3, col4, col5 = st.columns(5)
                    metrics = [
                        ("Profondeur", f"{depth} m"),
                        ("Em (Module)", f"{params.get('Em_MPa', '?')} MPa"),
                        ("Pl (Pression limite)", f"{params.get('Pl_MPa', '?')} MPa"),
                        ("Pf (Fluage)", f"{params.get('Pf_MPa', '?')} MPa"),
                        ("Em/Pl", f"{round(params['Em_MPa']/params['Pl_MPa'],1) if params.get('Em_MPa') and params.get('Pl_MPa') else '?'}"),
                    ]
                    for col, (label, val) in zip([col1,col2,col3,col4,col5], metrics):
                        with col:
                            st.markdown(f'<div class="metric-card"><h4>{label}</h4><p>{val}</p></div>', unsafe_allow_html=True)

                    st.markdown(f"**Classification du sol :** `{params.get('sol_type', 'N/A')}`")

                    if params.get("anomalies", 0) > 0:
                        st.markdown(f'<span class="anomaly-tag">⚠️ {params["anomalies"]} anomalie(s) détectée(s)</span>', unsafe_allow_html=True)
                    else:
                        st.markdown('<span class="ok-tag">✅ Aucune anomalie</span>', unsafe_allow_html=True)

                # Courbe P-V
                if not df_clean.empty:
                    fig = plot_courbe_pv(df_clean, depth, params)
                    st.plotly_chart(fig, use_container_width=True)

                    # Tableau corrigé
                    with st.expander("Données corrigées"):
                        st.dataframe(df_clean.round(3), use_container_width=True)

                # Métadonnées
                with st.expander("Métadonnées de l'essai"):
                    meta_df = pd.DataFrame([
                        {"Paramètre": k, "Valeur": str(v)}
                        for k, v in ess["meta"].items() if v is not None
                    ])
                    st.dataframe(meta_df, use_container_width=True)

    # ─── ONGLET 3 : Profil géotechnique ──────────────────────────────────────
    with tab_profil:
        st.subheader("🗻 Profil géotechnique complet")
        all_params = []
        progress = st.progress(0, text="Calcul en cours…")
        essais_list = list(data["essais"].items())

        for i, (name, ess) in enumerate(essais_list):
            depth = ess.get("depth_m", 0) or 0
            ref_et = ess["meta"].get("Ref étalonnage :")
            etalonnage = None
            for k, v in data["etalonnages"].items():
                if ref_et and ref_et in k:
                    etalonnage = v
                    break
            if etalonnage is None and data["etalonnages"]:
                etalonnage = list(data["etalonnages"].values())[0]

            df_c = clean_essai(ess, data["calibrage"], etalonnage)
            p = compute_pressiometre_params(df_c, depth)
            if p:
                p["name"] = name
                all_params.append(p)
            progress.progress((i + 1) / len(essais_list), text=f"Traitement : {name}")

        progress.empty()

        if all_params:
            # Tableau récap
            df_params = pd.DataFrame(all_params)
            display_cols = ["depth_m", "Em_MPa", "Pf_MPa", "Pl_MPa", "sol_type", "anomalies", "n_paliers"]
            display_cols = [c for c in display_cols if c in df_params.columns]
            st.dataframe(df_params[display_cols].sort_values("depth_m").round(3),
                         use_container_width=True)

            # Profil graphique
            fig_p = plot_profil(all_params)
            st.plotly_chart(fig_p, use_container_width=True)

            # Export CSV
            csv = df_params.to_csv(index=False).encode("utf-8")
            st.download_button(
                "⬇️ Télécharger résultats (CSV)",
                data=csv,
                file_name=f"pressiometre_resultats_{datetime.date.today()}.csv",
                mime="text/csv",
            )
        else:
            st.warning("Aucun paramètre calculable. Vérifiez les données.")

    # ─── ONGLET 4 : IA KIBALI ────────────────────────────────────────────────
    with tab_ia:
        st.subheader("🤖 Assistant KIBALI — Expert Géotechnique")

        if not use_ia or model is None:
            st.info("Activez l'IA KIBALI dans les paramètres (barre gauche) pour utiliser cette fonctionnalité.")
            st.markdown("""
            **Exemples de questions que KIBALI peut répondre :**
            - *Interprète les résultats de cet essai pressiométrique à 12 m*
            - *Quelles anomalies observes-tu dans ces données ?*
            - *Quel type de fondation recommandes-tu pour ce profil ?*
            - *Explique la différence entre Pl et Pf*
            - *Est-ce que ce sol est adapté pour une fondation superficielle ?*
            """)
            return

        # Contexte automatique depuis le fichier chargé
        context_str = ""
        if data and data["essais"]:
            lines = [f"Sondage : {source_name}", f"Nombre d'essais : {len(data['essais'])}"]
            for name, ess in list(data["essais"].items())[:5]:
                depth = ess.get("depth_m", "?")
                n = len([m for m in ess["mesures"] if m["P60_MPa"]])
                lines.append(f"  - {name} : profondeur={depth}m, {n} paliers")
            context_str = "\n".join(lines)

        # Historique de chat
        if "chat_history" not in st.session_state:
            st.session_state.chat_history = []

        # Affichage historique
        for msg in st.session_state.chat_history:
            with st.chat_message(msg["role"], avatar="🌍" if msg["role"] == "assistant" else "👤"):
                st.markdown(msg["content"])

        # Questions rapides
        st.markdown("**Questions rapides :**")
        qcols = st.columns(3)
        quick_qs = [
            "Interprète ce profil géotechnique",
            "Y a-t-il des anomalies dans les données ?",
            "Quel type de fondation recommandes-tu ?",
        ]
        for i, (qcol, qq) in enumerate(zip(qcols, quick_qs)):
            if qcol.button(qq, key=f"q_{i}", use_container_width=True):
                st.session_state["quick_q"] = qq

        # Zone de saisie
        user_input = st.chat_input("Posez une question à KIBALI…")
        if not user_input and st.session_state.get("quick_q"):
            user_input = st.session_state.pop("quick_q")

        if user_input:
            st.session_state.chat_history.append({"role": "user", "content": user_input})
            with st.chat_message("user", avatar="👤"):
                st.markdown(user_input)

            with st.chat_message("assistant", avatar="🌍"):
                with st.spinner("KIBALI réfléchit…"):
                    response = ask_kibali(model, tokenizer, user_input, context_str)
                st.markdown(response)

            st.session_state.chat_history.append({"role": "assistant", "content": response})

        if st.session_state.chat_history:
            if st.button("🗑️ Effacer la conversation"):
                st.session_state.chat_history = []
                st.rerun()


if __name__ == "__main__":
    main()
