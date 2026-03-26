"""
Parser Excel → ParsedFile
Gère les fichiers pressiométriques Ménard (NF P 94-110)
"""
from __future__ import annotations
import io
import re
import datetime
from pathlib import Path
from typing import Union, Optional

import openpyxl

from .models import (
    EssaiMeta, EssaiRaw, ParsedFile, RawMesure
)

# Correspondance des clés Excel → champs EssaiMeta
_META_MAP = {
    "projet":           ["projet :"],
    "localisation":     ["localisation :"],
    "ref_sondage":      ["ref sondage :"],
    "ref_essai":        ["ref essai :"],
    "ref_sonde":        ["ref sonde :"],
    "ref_etalonnage":   ["ref étalonnage :", "ref etalonnage :"],
    "ref_calibrage":    ["ref calibrage :"],
    "passe_forage":     ["passe de forage (m) :"],
    "technique":        ["tech. utilisée:", "tech. utilisee:"],
    "outil_forage":     ["outil de forage :"],
    "pression_diff_bar":["pression diff (bar) :"],
    "type_tubulure":    ["type de tubulure :"],
    "date":             ["date :"],
    "profondeur_m":     ["prof. de l'essai (m) :", "prof. de l essai (m) :"],
}


def _norm(s: str) -> str:
    return s.lower().strip().replace("'", " ").replace("'", " ")


def _parse_meta(raw_key: str, raw_val) -> tuple[Optional[str], Optional[object]]:
    key_norm = _norm(str(raw_key))
    for field, aliases in _META_MAP.items():
        for alias in aliases:
            if key_norm == _norm(alias):
                return field, raw_val
    return None, None


def _parse_depth(val) -> Optional[float]:
    if val is None:
        return None
    try:
        return float(str(val).replace(",", ".").replace("m", "").replace("M", "").strip())
    except (ValueError, TypeError):
        return None


def _extract_depth_from_name(sheet_name: str) -> Optional[float]:
    m = re.search(r"(\d+(?:[.,]\d+)?)\s*m", sheet_name, re.IGNORECASE)
    if m:
        return float(m.group(1).replace(",", "."))
    return None


def parse_sheet(ws) -> EssaiRaw:
    """Parse une feuille Excel → EssaiRaw."""
    meta_dict: dict = {}
    mesures: list[RawMesure] = []
    in_data = False
    sheet_name = ws.title.strip()

    for row in ws.iter_rows(values_only=True):
        vals = list(row)
        first = vals[0]

        # Détection début des données brutes
        if isinstance(first, str) and "donnees brutes" in _norm(first):
            in_data = True
            continue

        if not in_data:
            if first is None:
                continue
            key_str = str(first).strip()
            val = vals[1] if len(vals) > 1 else None
            field, parsed_val = _parse_meta(key_str, val)
            if field:
                # Conversion spéciale pour la date
                if field == "date" and isinstance(parsed_val, datetime.datetime):
                    meta_dict[field] = parsed_val.strftime("%Y-%m-%d")
                elif field == "profondeur_m":
                    meta_dict[field] = _parse_depth(parsed_val)
                elif field == "pression_diff_bar":
                    try:
                        meta_dict[field] = float(str(parsed_val).replace(",", "."))
                    except (ValueError, TypeError):
                        meta_dict[field] = 0.0
                else:
                    meta_dict[field] = str(parsed_val).strip() if parsed_val is not None else None
        else:
            # Ligne d'en-tête des données (Numéro du palier, V30, P60, V60…)
            if isinstance(first, str):
                continue
            # Ligne de mesure : first = numéro de palier (int/float)
            if first is None:
                continue
            try:
                palier = int(first)
            except (ValueError, TypeError):
                continue

            v30 = vals[2] if len(vals) > 2 else None
            p60 = vals[3] if len(vals) > 3 else None
            v60 = vals[4] if len(vals) > 4 else None

            # Ignorer les string sentinelles comme '_-10'
            if isinstance(v30, str):
                v30 = None
            if isinstance(v60, str):
                v60 = None
            if isinstance(p60, str):
                try:
                    p60 = float(p60.replace(",", "."))
                except (ValueError, AttributeError):
                    p60 = None

            # Filtrage : conserver si au moins P60 est présent
            if p60 is not None or v30 is not None or v60 is not None:
                try:
                    v30_f = float(v30) if v30 is not None else None
                    p60_f = float(str(p60).replace(",", ".")) if p60 is not None else None
                    v60_f = float(v60) if v60 is not None else None
                except (ValueError, TypeError):
                    continue
                mesures.append(RawMesure(
                    palier=palier,
                    V30_cm3=v30_f,
                    P60_MPa=p60_f,
                    V60_cm3=v60_f,
                ))

    # Construire EssaiMeta
    meta = EssaiMeta(**{k: v for k, v in meta_dict.items() if k in EssaiMeta.model_fields})

    # Profondeur depuis métadonnées ou nom de feuille
    depth_m = meta.profondeur_m
    if depth_m is None:
        depth_m = _extract_depth_from_name(sheet_name)

    # Déterminer si c'est calibrage / étalonnage
    name_low = sheet_name.lower()
    is_cal = any(k in name_low for k in ["calibr"])
    is_et  = any(k in name_low for k in ["etalon", "étalon"])

    return EssaiRaw(
        sheet_name=sheet_name,
        meta=meta,
        mesures=mesures,
        depth_m=depth_m,
        is_calibrage=is_cal,
        is_etalonnage=is_et,
    )


def load_excel(source: Union[str, Path, bytes, io.BytesIO],
               filename: Optional[str] = None) -> ParsedFile:
    """
    Charge un fichier Excel pressiométrique.
    source : chemin fichier, bytes bruts, ou BytesIO.
    Retourne ParsedFile avec essais, calibrage, étalonnages.
    """
    if isinstance(source, (str, Path)):
        wb = openpyxl.load_workbook(str(source), data_only=True)
        filename = filename or Path(source).name
    elif isinstance(source, (bytes, bytearray)):
        wb = openpyxl.load_workbook(io.BytesIO(source), data_only=True)
        filename = filename or "upload.xlsx"
    else:
        # io.BytesIO ou compatible
        source.seek(0)
        wb = openpyxl.load_workbook(source, data_only=True)
        filename = filename or "upload.xlsx"

    essais: dict[str, EssaiRaw] = {}
    calibrage: Optional[EssaiRaw] = None
    etalonnages: dict[str, EssaiRaw] = {}

    for shname in wb.sheetnames:
        ws = wb[shname]
        parsed = parse_sheet(ws)

        if parsed.is_calibrage:
            calibrage = parsed
        elif parsed.is_etalonnage:
            etalonnages[shname.strip()] = parsed
        else:
            # Garder seulement les feuilles avec des mesures
            if parsed.mesures:
                essais[shname.strip()] = parsed

    return ParsedFile(
        filename=filename,
        essais=essais,
        calibrage=calibrage,
        etalonnages=etalonnages,
    )
